"""
Extract plain text from uploaded files.
Supports: PDF, DOCX, TXT, MD, CSV, XLSX, PY/JS/code, images (caption).
Falls back to raw read for unknown types.
"""
import csv
import io
from pathlib import Path


def extract_text(file_path: str, filename: str) -> str:
    ext  = Path(filename).suffix.lower()
    path = Path(file_path)

    try:
        if ext == ".pdf":
            return _pdf(path)
        if ext in (".docx", ".doc"):
            return _docx(path)
        if ext in (".xlsx", ".xls"):
            return _xlsx(path)
        if ext == ".csv":
            return _csv(path)
        if ext in (".txt", ".md", ".rst", ".log"):
            return path.read_text(errors="ignore")
        if ext in (".py", ".js", ".ts", ".jsx", ".tsx", ".java",
                   ".cpp", ".c", ".go", ".rs", ".rb", ".php",
                   ".html", ".css", ".json", ".xml", ".yaml", ".yml", ".sh", ".sql"):
            return f"[Code file: {filename}]\n\n" + path.read_text(errors="ignore")
        if ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            return f"[Image file: {filename}] — visual content, no text extracted."
        if ext in (".mp4", ".mov", ".avi", ".mkv"):
            return f"[Video file: {filename}] — cannot extract text from video."
        # Fallback: try reading as text
        return path.read_text(errors="ignore")
    except Exception as e:
        return f"[Could not extract text from {filename}: {e}]"


def _pdf(path: Path) -> str:
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(str(path))
        return "\n\n".join(
            page.extract_text() or "" for page in reader.pages
        )
    except Exception as e:
        return f"[PDF extraction error: {e}]"


def _docx(path: Path) -> str:
    try:
        from docx import Document
        doc = Document(str(path))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        return f"[DOCX extraction error: {e}]"


def _xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb   = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
        return "\n".join(rows)
    except Exception as e:
        return f"[XLSX extraction error: {e}]"


def _csv(path: Path) -> str:
    try:
        text = path.read_text(errors="ignore")
        reader = csv.reader(io.StringIO(text))
        return "\n".join("\t".join(row) for row in reader)
    except Exception as e:
        return f"[CSV extraction error: {e}]"


def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 150) -> list[str]:
    """Split text into overlapping chunks."""
    if not text.strip():
        return []
    words  = text.split()
    chunks = []
    step   = max(1, chunk_size - overlap)
    for i in range(0, len(words), step):
        chunk = " ".join(words[i : i + chunk_size])
        if chunk.strip():
            chunks.append(chunk)
    return chunks
